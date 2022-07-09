from selenium import webdriver
import time
from selenium.webdriver.common.by import By
import pytest
import openpyxl


#configure workbook path
wb = openpyxl.load_workbook("C:\\Users\\jagad\\Dropbox\\PC\\Desktop\\Test_data.xlsx")
#get active sheet
sheet = wb.active
rows = sheet.max_row
cols = sheet.max_column
list=[]
#get cell address within active sheet:
for r in range(2, rows + 1):
    username = sheet.cell(r, 1).value
    password = sheet.cell(r, 2).value
    tuple = (username, password)
    list.append(tuple)
    #print(list)

@pytest.mark.parametrize("username,password",list)
def test_guru99(username,password):
    global driver
    driver = webdriver.Chrome(executable_path="C:\\Users\\jagad\\Dropbox\\PC\Desktop\\automation_framework\\dependencies\\drivers\\chrome\\chromedriver.exe")
    driver.implicitly_wait(5)
    driver.maximize_window()
    driver.get("http://www.demo.guru99.com/V4/")
    driver.find_element(By.NAME, "uid").send_keys(username)
    driver.find_element(By.NAME, "password").send_keys(password)
    driver.find_element(By.NAME, "btnLogin").click()
    time.sleep(5)
    driver.quit()
