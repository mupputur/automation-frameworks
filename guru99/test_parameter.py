from selenium import webdriver
import time
from selenium.webdriver.common.by import By
import pytest
import openpyxl
from Utils import url,user_name,pass_word,bot_ton,path,xl_path

def read_data():
    #configure workbook path
    wb = openpyxl.load_workbook(xl_path)
    #get active sheet
    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column
    list=[]
    #get cell address within active sheet:
    for r in range(2, rows + 1):
        username = sheet.cell(r, 1).value
        password = sheet.cell(r, 2).value
        tuple = (username, password)
        list.append(tuple)
        #print(list)
        return list

@pytest.mark.parametrize("username,password",list)
def test_guru99(username,password):
    global driver
    driver = webdriver.Chrome(path)
    driver.implicitly_wait(5)
    driver.maximize_window()
    driver.get(url)
    driver.find_element(By.NAME,user_name ).send_keys(username)
    driver.find_element(By.NAME,pass_word ).send_keys(password)
    driver.find_element(By.NAME, bot_ton).click()
    time.sleep(5)
    driver.quit()
